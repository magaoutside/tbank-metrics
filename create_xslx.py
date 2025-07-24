import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font

def process_dialogs(csv_path):
    dialogs = []
    current_dialog = []
    
    with open(csv_path, 'r', encoding='utf-8') as file:
        csv_reader = csv.reader(file)
        next(csv_reader)  # Пропускаем заголовок
        
        for row in csv_reader:
            # Пропускаем пустые строки
            if not row or all(field.strip() == '' for field in row):
                if current_dialog:
                    dialogs.append(current_dialog)
                    current_dialog = []
                continue
                
            # Берем текст и инструмент из столбцов
            text = row[0].strip() if len(row) > 0 and row[0].strip() else ''
            tool = row[1].strip() if len(row) > 1 and row[1].strip() else ''
            
            if text:
                current_dialog.append((text, tool))
        
        # Добавляем последний диалог если есть
        if current_dialog:
            dialogs.append(current_dialog)
    
    return dialogs

def create_xlsx(dialog, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dialog"
    
    # Создаем заголовки
    headers = ["Role", "Content", "Needed_Tool"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    row_num = 2
    for text, tool in dialog:
        # Строка пользователя
        ws.cell(row=row_num, column=1, value="user")
        ws.cell(row=row_num, column=2, value=text)
        ws.cell(row=row_num, column=3, value=tool)
        row_num += 1
        
        # Строка бота
        ws.cell(row=row_num, column=1, value="bot")
        ws.cell(row=row_num, column=2, value="«тут должен быть текст»")
        ws.cell(row=row_num, column=3, value="")  # Пустой для бота
        row_num += 1
    
    # Автонастройка ширины колонок
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    wb.save(filename)

# Основной код
if __name__ == "__main__":
    csv_file = "dialog.csv"
    output_dir = "dialogs"
    
    # Создаем папку если ее нет
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Создана папка: {output_dir}")
    
    # Обрабатываем диалоги
    dialogs = process_dialogs(csv_file)
    
    # Создаем XLSX файлы для каждого диалога
    for i, dialog in enumerate(dialogs, start=1):
        filename = os.path.join(output_dir, f"{i}.xlsx")
        create_xlsx(dialog, filename)
        print(f"Создан файл: {filename} с {len(dialog)} сообщениями пользователя")
    
    print(f"\nВсего создано файлов: {len(dialogs)}")
    print(f"Все файлы сохранены в папке: {output_dir}")